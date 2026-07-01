#!/usr/bin/env python3
from __future__ import annotations

import json
import math
import os
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
RAW_DATA_DIR = ROOT / "data" / "raw"
DEFAULT_TARGETS_XLSX = RAW_DATA_DIR / "副本1a8083ce4a7ced5847024a560e3ed22b.xlsx"
DEFAULT_TRACKS_XLSX = RAW_DATA_DIR / "副本0cb4b68fa1a67179a0368da8eb82dff6.xlsx"
OUTPUT = ROOT / "src" / "data" / "seasatsPayload.json"

MONITORED_AREAS = [
    {"id": "bahrain-gulf-test", "name": "巴林港区/波斯湾测试点", "center": {"lon": 50.608, "lat": 26.205}, "radiusNm": 3},
    {"id": "san-diego-naval-area", "name": "圣迭戈军港周边", "center": {"lon": -117.24, "lat": 32.77}, "radiusNm": 15},
    {"id": "south-china-sea-low-speed", "name": "南海低速活动区", "center": {"lon": 120.985, "lat": 14.562}, "radiusNm": 10},
    {"id": "taiwan-southwest-offshore", "name": "台湾西南近海活动区", "center": {"lon": 120.28, "lat": 22.6}, "radiusNm": 10},
]

PARAMETERS = {
    "lowSpeedMaxKn": 3,
    "lowSpeedDurationMinutes": 10,
    "repeatedPathRatio": 3,
    "aisGapWarningMinutes": 30,
    "aisGapCriticalMinutes": 360,
    "segmentGapMinutes": 360,
    "segmentJumpNm": 50,
}


def clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def to_float(value: Any) -> float | None:
    value = clean(value)
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def to_int(value: Any) -> int | None:
    number = to_float(value)
    return None if number is None else int(number)


def to_text(value: Any) -> str | None:
    value = clean(value)
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def to_iso(value: Any) -> str | None:
    value = clean(value)
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%dT%H:%M:%SZ")
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%dT%H:%M:%SZ")
        except ValueError:
            pass
    return text


def rows_from_workbook(path: Path) -> tuple[str, list[dict[str, Any]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(clean(c)) for c in rows[0]]
    records = []
    for source_row, row in enumerate(rows[1:], start=2):
        record = {headers[i]: clean(row[i]) if i < len(row) else None for i in range(len(headers))}
        if any(value is not None for value in record.values()):
            record["_sourceRow"] = source_row
            records.append(record)
    return ws.title, records


def build_targets(path: Path) -> tuple[str, list[dict[str, Any]]]:
    sheet, records = rows_from_workbook(path)
    targets = []
    for row in records:
        raw_speed = to_float(row.get("速度（除10）"))
        raw_cog = to_float(row.get("cog（除10）"))
        targets.append({
            "mmsi": to_text(row.get("mmsi")),
            "name": to_text(row.get("NAME")),
            "latestTime": to_iso(row.get("最后位置时间")),
            "lon": to_float(row.get("longitude")),
            "lat": to_float(row.get("latitude")),
            "speedRawDiv10": raw_speed,
            "speedKn": None if raw_speed is None else round(raw_speed / 10, 3),
            "courseDeg": None if raw_cog is None else round(raw_cog / 10, 1),
            "length": to_int(row.get("LENGTH")),
            "width": to_int(row.get("width")),
            "sourceRow": row["_sourceRow"],
        })
    return sheet, targets


def build_track_points(path: Path) -> tuple[str, list[dict[str, Any]]]:
    sheet, records = rows_from_workbook(path)
    points = []
    for row in records:
        points.append({
            "id": f"{to_text(row.get('mmsi'))}-{row['_sourceRow']}",
            "mmsi": to_text(row.get("mmsi")),
            "time": to_iso(row.get("quire_time")),
            "lon": to_float(row.get("lng")),
            "lat": to_float(row.get("lat")),
            "speedKn": to_float(row.get("speed")),
            "speedKmh": to_float(row.get("speed_km")),
            "rateOfTurn": to_float(row.get("rate_of_turn")),
            "orientation": to_float(row.get("orientation")),
            "heading": to_float(row.get("heading")),
            "navStatus": to_text(row.get("nav_status")),
            "aisSourceType": to_text(row.get("ais_source_type")),
            "seaName": to_text(row.get("sea_name")),
            "provider": to_text(row.get("provider")),
            "confidence": to_float(row.get("confidence")),
            "sourceRow": row["_sourceRow"],
        })
    points.sort(key=lambda p: (p["mmsi"] or "", p["time"] or ""))
    return sheet, points


def resolve_input_path(value: str | None, default: Path) -> Path:
    if not value:
        return default
    path = Path(value)
    return path if path.is_absolute() else ROOT / path


def project_relative_path(path: Path) -> str:
    try:
        return str(path.resolve().relative_to(ROOT))
    except ValueError:
        return str(path)


def main() -> None:
    targets_path = resolve_input_path(os.environ.get("SEASATS_TARGETS_XLSX"), DEFAULT_TARGETS_XLSX)
    tracks_path = resolve_input_path(os.environ.get("SEASATS_TRACKS_XLSX"), DEFAULT_TRACKS_XLSX)
    target_sheet, targets = build_targets(targets_path)
    track_sheet, track_points = build_track_points(tracks_path)
    times = [p["time"] for p in track_points if p.get("time")]
    target_times = [t["latestTime"] for t in targets if t.get("latestTime")]
    payload = {
        "metadata": {
            "generatedAt": datetime.now(UTC).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "targetWorkbook": {"fileName": targets_path.name, "path": project_relative_path(targets_path), "sheet": target_sheet},
            "trackWorkbook": {"fileName": tracks_path.name, "path": project_relative_path(tracks_path), "sheet": track_sheet},
            "targetCount": len(targets),
            "trackPointCount": len(track_points),
            "trackMmsiCount": len({p["mmsi"] for p in track_points if p.get("mmsi")}),
            "dataWindow": {
                "start": min(times) if times else None,
                "end": max(times) if times else None,
                "latestPositionStart": min(target_times) if target_times else None,
                "latestPositionEnd": max(target_times) if target_times else None,
            },
        },
        "parameters": PARAMETERS,
        "monitoredAreas": MONITORED_AREAS,
        "targets": targets,
        "trackPoints": track_points,
    }
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "output": str(OUTPUT),
        "targets": len(targets),
        "trackPoints": len(track_points),
        "trackMmsiCount": payload["metadata"]["trackMmsiCount"],
        "dataWindow": payload["metadata"]["dataWindow"],
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
